# import os
# from openpyxl import Workbook, load_workbook

# # Writing data to an Excel file
# def write_to_excel(filename):
#     # Create a new Workbook and select the active worksheet
#     workbook = Workbook()
#     sheet = workbook.active
    
#     # Custom data to write
#     data = [
#          ["Name", "Age", "City", "Country", "Semester", "Class", "Group"],
#         ["Ashish", 18, "Mumbai", "India", "1st", "B.tech", "A"],
#         ["Saheb", 15, "Los Angeles", "USA", "2nd", "MBA", "B"],
#         ["Rahul", 12, "Toronto", "Canada", "3rd", "B.Com", "A"],
#         ["Dilip", 18, "Beijing", "China", "1st", "B.Sc", "C"],
#         ["Rohit", 35, "Delhi", "India", "2nd", "MBA", "B"],
#         ["Sandeep", 22, "Bettiah", "India", "3rd", "B.Com", "A"],
#         ["John", 20, "New York", "USA", "1st", "B.Sc", "D"],
#         ["Michael", 23, "London", "UK", "2nd", "B.Com", "A"],
#         ["Emily", 21, "Sydney", "Australia", "1st", "B.Sc", "C"],
#         ["Jessica", 24, "Tokyo", "Japan", "3rd", "MBA", "B"],
#         ["David", 19, "Berlin", "Germany", "1st", "B.Com", "D"],
#         ["Emma", 22, "Paris", "France", "2nd", "B.Sc", "A"],
#         ["Oliver", 25, "Moscow", "Russia", "3rd", "MBA", "C"],
#         ["Charlotte", 20, "Dubai", "UAE", "1st", "B.Com", "B"],
#         ["Liam", 26, "Singapore", "Singapore", "2nd", "B.Sc", "A"],
#         ["Sophia", 18, "Seoul", "South Korea", "1st", "B.Com", "C"],
#         ["Isabella", 22, "Barcelona", "Spain", "3rd", "MBA", "D"],
#         ["Lucas", 21, "Rome", "Italy", "2nd", "B.Sc", "B"],
#         ["Ava", 19, "Amsterdam", "Netherlands", "1st", "B.Com", "A"],
#         ["Mason", 23, "Cairo", "Egypt", "2nd", "B.Sc", "C"],
#         ["Mia", 24, "Hanoi", "Vietnam", "3rd", "MBA", "D"],
#         ["Noah", 20, "Bangkok", "Thailand", "1st", "B.Com", "B"],
#         ["Olivia", 25, "Lagos", "Nigeria", "2nd", "B.Sc", "A"],
#         ["Ethan", 19, "Buenos Aires", "Argentina", "1st", "B.Com", "C"]


#     ]
    
#     # Write data to the worksheet
#     for row in data:
#         sheet.append(row)
    
#     # Save the workbook
#     workbook.save(filename)
#     print(f"Data written to {filename} successfully.")
    
    
#     # Open the Excel file after writing (Windows)
#     os.startfile(filename)

# # Reading data from an Excel file
# def read_from_excel(filename):
#     # Load the workbook and select the active worksheet
#     workbook = load_workbook(filename)
#     sheet = workbook.active
    
#     # Read and print the data
#     print(f"Data read from {filename}:")
#     for row in sheet.iter_rows(values_only=True):
#         print(row)

# # Example usage
# excel_file = "example.xlsx"
# write_to_excel(excel_file)
# read_from_excel(excel_file)






# import csv
# from openpyxl import Workbook
# from openpyxl import load_workbook
# def write_to_excel(data, file_name):
#     workbook = Workbook()
#     sheet = workbook.active

#     for row in data:
#         sheet.append(row)

#     workbook.save(file_name)
#     print("Data successfully written to {file_name}")
# def write_to_csv(data, file_name):
    
    
#     with open(file_name, 'w', newline='') as csvfile:
#         writer = csv.writer(csvfile)
#         writer.writerows(data)
#     print(f"Data successfully written to {file_name}")
    
    
# def read_from_excel(file_name):
#     workbook = load_workbook(file_name)
#     sheet = workbook.active
#     data = []
#     for row in sheet.iter_rows(values_only=True):
#         data.append(row)
#     return data


# def read_from_csv(file_name):
#     with open(file_name, 'r') as csvfile:
#         reader = csv.reader(csvfile)
#         data = list(reader)
#     return data


# # Example usage
# data = [
#     ["Name", "Age", "City"],
#     ["Ashish", 18, "New York"],
#     ["Saheb", 15, "Los Angeles"],
#     ["Rahul", 12, "Chicago"]
# ]
# # Write data to Excel file
# write_to_excel (data, 'data.xlsx')
# # Write data to CSV file
# write_to_csv(data, 'data.csv')
# # Read data from Excel file
# excel_data = read_from_excel('data.xlsx')
# print("Data read from Excel:")
# for row in excel_data:
#     print(row)
# # Read data from CSV file
# csv_data = read_from_csv('data.csv')
# print("Data read from CSV:")
# for row in csv_data:
#     print(row)







import csv 
from openpyxl import Workbook, load_workbook
# Writing data to a CSV file
def write_to_csv(filename):
    # Custom data to write
    data = [
        ["Name", "Age", "City"],
        ["Ashish", 18, "Bettih"],
        ["Rahul", 15, "Motihari"],
        ["Ankit", 22, "Patna"]
    ]
# Write data to CSV
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(data)
    print(f"Data written to {filename} successfully.")

# Reading data from a CSV file
def read_from_csv(filename):
    # Read data from CSV
    with open(filename, mode='r') as file:
        reader = csv.reader(file)
        print(f"Data read from {filename}:")
        for row in reader:
            print(row)

# Example usage
csv_file = "myfile.csv"
write_to_csv(csv_file)
read_from_csv(csv_file)










